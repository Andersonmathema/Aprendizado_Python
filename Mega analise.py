from openpyxl import*
from openpyxl.utils import get_column_letter
from random import randint
from time import sleep
lista = list()
listar = list()
jogos = list()
jogosr = list()
arq = load_workbook('F:\OneDrive\Projetos Python\Megasena.xlsx')
sheet = arq.active
lmax = sheet.max_row
for row in range(2, lmax):
    for col in range(1, 8):
        char = get_column_letter(col) #Pega as letras das colunas: A, B, C...
        lin = sheet[char + str(row)].value #Mostra o valor em colunas das letras+range, ficando A1, B1, ...
        lista.append(lin)
    jogos.append(lista[:])
    lista.clear()
nump = 0
'''for item in jogos:
    contagem = item.count(13)
    if contagem == True:
        nump += 1 # este nump conta quantos numeros indicados (no caso 13) apareceram na lista jogos
print(nump)'''
'''print(jogos[-6:])
for item in jogos[-6:]:
    if 2 in item:
        print('ok')
    else:
        print('no')'''
quant = int(input('Quantos jogos você quer sortear?: '))
tot = 1
som = 0
while tot <= quant:
    cont = 0
    while True:
        num = randint(1, 60)
        if num not in listar:
            listar.append(num)
            cont += 1
        if cont >= 6:
            break
    listar.sort()
    jogosr.append(listar[:])
    listar.clear()
    tot += 1
print('-='*3, f'SORTEANDO {quant} JOGOS', '-='*3)
for i, l in enumerate(jogosr):
    print(f'Jogo {i+1}: {l}')
    sleep(1)
print('-=-=-=-=-=-=< BOA SORTE >-=-=-=-=-=-=')

