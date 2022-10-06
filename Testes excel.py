from openpyxl import*
wb = Workbook()
sh = wb.active
'''sh['A1'] = 'Programe'
sh['B1'] = 'Python'
wb.save(filename = 'F:\OneDrive\Projetos Python\Arquivopy.xlsx') #salva ou cria planilha
del wb'''
wb = load_workbook('F:\OneDrive\Projetos Python\Arquivopy.xlsx')
print(wb.sheetnames)
sh = wb['Teste']
sh2 = wb['Sheet']
'''cel = sh2.cell(row=1, column=1).value 
print(cel)
cel = 'PROGRAME'
print(cel)
cel = sh2.cell(row=1, column=1, value = 'PROG').value #Atribuir valores a celulas especificas
print(cel)'''
f = sh2['A1:B2']
for row in sh2.iter_rows():
    row[0].value = 5
    row[1].value = 6
    print(row[0].value)
    print(row[1].value)
sh2['C2'] = '=AVERAGE(A1:B2)'
print(sh2['C2'].value)
wb.save(filename = 'F:\OneDrive\Projetos Python\Arquivopy.xlsx')