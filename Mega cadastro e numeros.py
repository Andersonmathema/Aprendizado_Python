from openpyxl import*
lista = list()
jogos = list()
arq = load_workbook('F:\OneDrive\Projetos Python\Megasena.xlsx')
sheet = arq.active
cont = 0
cont2 = 0
concurso = 0
lwrite = sheet.max_row + 1 #Procura a última linha e adiciona 1 para direcionar a escrita para a próxima linha
lin = f'A{lwrite}:G{lwrite}'
print(lin)
print(sheet.max_row)
while cont2 == 0:
    concurso = int(input('Digite o número do concurso: '))
    cont2 += 1
while True:
    num = int(input(f'Digite um número: '))
    if num not in lista:
        lista.append(num)
        cont += 1
    if cont >= 6:
        break
lista.sort()
lista.insert(0, concurso)
jogos.append(lista[:])
for row in sheet[lin]:
    for index, cell in enumerate(row):
        cell.value = lista[index]
arq.save('F:\OneDrive\Projetos Python\Megasena.xlsx')
